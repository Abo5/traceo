"""v2 addendum quality gates: public API keys + CI gate, X-API-Key alt-auth on the
public surface, webhooks (CRUD/test/fire — no external network, httpx monkeypatched),
schedules + scheduler tick, Xray/Jira exports, org data export, feature reference."""
import hashlib
import hmac
import json
import time
from datetime import datetime, timedelta, timezone

import pytest
from conftest import (add_requirement, confirm_requirement, import_spec, items_of,
                      poll_job)

DEAD_BASE_URL = "http://127.0.0.1:9"  # connection refused instantly -> errored results


# ------------------------------------------------------------------ helpers

def seed_approved_cases(client, headers, pid):
    """Manual requirement -> confirm -> spec import -> generate -> approve all."""
    rid = add_requirement(
        client, headers, pid, "REQ-001",
        "The customer phone number must start with 05 and be exactly 10 digits "
        "when creating a customer through POST /customers.",
        criteria=["reject any phone that does not match 05XXXXXXXX with a 422 (invalid phone rejected)",
                  "accept a valid phone such as 0512345678 (valid phone accepted for customers)"])
    confirm_requirement(client, headers, rid)
    import_spec(client, headers, pid)
    r = client.post(f"/v1/projects/{pid}/generate", json={"depth": "smoke"},
                    headers=headers)
    assert r.status_code in (200, 202), r.text
    job = poll_job(client, headers, r.json()["job_id"])
    assert (job.get("result") or {}).get("generated", 0) > 0
    drafts = items_of(client.get(f"/v1/projects/{pid}/test-cases",
                                 params={"state": "draft"}, headers=headers).json())
    assert drafts
    r = client.post("/v1/test-cases/bulk",
                    json={"ids": [t["id"] for t in drafts], "action": "approve"},
                    headers=headers)
    assert r.status_code in (200, 201, 204), r.text
    return rid


def make_env(client, headers, pid, base_url=DEAD_BASE_URL):
    r = client.post(f"/v1/projects/{pid}/environments",
                    json={"name": "staging", "base_url": base_url,
                          "auth_type": "none"}, headers=headers)
    assert r.status_code == 201, r.text
    return r.json()["id"]


def make_api_key(client, headers, name="ci-key"):
    r = client.post("/v1/api-keys", json={"name": name}, headers=headers)
    assert r.status_code == 201, r.text
    return r.json()


def wait_run_terminal(client, headers, run_id, timeout=30.0):
    deadline = time.time() + timeout
    while time.time() < deadline:
        r = client.get(f"/v1/runs/{run_id}", headers=headers)
        assert r.status_code == 200, r.text
        run = r.json()
        if run["state"] in ("completed", "cancelled", "aborted"):
            return run
        time.sleep(0.2)
    raise AssertionError(f"run {run_id} did not reach a terminal state")


class _FakeResp:
    def __init__(self, status_code=200):
        self.status_code = status_code


@pytest.fixture()
def webhook_net(monkeypatch):
    """No external network: bypass the SSRF DNS guard and capture httpx.post."""
    calls = []

    def fake_post(url, content=b"", headers=None, timeout=None, **kw):
        calls.append({"url": url, "content": content, "headers": headers or {}})
        return _FakeResp(200)

    monkeypatch.setattr("app.modules.integrations._assert_public_host",
                        lambda hostname: None)
    monkeypatch.setattr("app.modules.integrations.httpx.post", fake_post)
    return calls


# ------------------------------------------------------------------ API keys + gate

def test_api_key_lifecycle_and_gate_auth(client, register_org, create_project):
    headers = register_org()
    pid = create_project(headers)

    created = make_api_key(client, headers)
    key = created["key"]
    assert key.startswith("trc_") and len(key) == 44  # trc_ + 40 hex
    assert created["prefix"] == key[:8]

    # list never exposes the full key
    listed = client.get("/v1/api-keys", headers=headers).json()
    assert len(listed) == 1
    assert "key" not in listed[0] and listed[0]["prefix"] == key[:8]
    assert listed[0]["revoked"] is False

    # gate works with X-API-Key AND with Bearer; empty project fails coverage
    for auth in ({"X-API-Key": key}, headers):
        r = client.get(f"/v1/projects/{pid}/gate", headers=auth)
        assert r.status_code == 200, r.text
        gate = r.json()
        assert gate["pass"] is False
        assert any(b["check"] == "min_coverage" for b in gate["breaches"])

    # last_used_at recorded after use
    listed = client.get("/v1/api-keys", headers=headers).json()
    assert listed[0]["last_used_at"]

    # ?exit=1 -> 412 for `curl -f` pipelines
    r = client.get(f"/v1/projects/{pid}/gate", params={"exit": 1},
                   headers={"X-API-Key": key})
    assert r.status_code == 412
    assert r.json()["detail"]["code"] == "gate_failed"

    # unknown key and revoked key -> 401
    r = client.get(f"/v1/projects/{pid}/gate", headers={"X-API-Key": "trc_" + "0" * 40})
    assert r.status_code == 401
    r = client.post(f"/v1/api-keys/{created['id']}/revoke", headers=headers)
    assert r.status_code == 200 and r.json()["revoked"] is True
    r = client.get(f"/v1/projects/{pid}/gate", headers={"X-API-Key": key})
    assert r.status_code == 401
    assert r.json()["detail"]["code"] == "invalid_api_key"


def test_gate_thresholds_with_seeded_flow(client, register_org, create_project):
    headers = register_org()
    pid = create_project(headers)
    seed_approved_cases(client, headers, pid)

    # 1 confirmed requirement, fully covered, no runs -> gate passes
    gate = client.get(f"/v1/projects/{pid}/gate", headers=headers).json()
    assert gate["pass"] is True and gate["breaches"] == []
    assert gate["coverage_pct"] == 100.0
    assert gate["open_defects"] == {"total": 0, "critical": 0}
    assert gate["latest_run"] is None

    # add an uncovered confirmed requirement -> coverage 50% -> min_coverage breach
    rid2 = add_requirement(client, headers, pid, "REQ-XXX",
                           "A requirement with no test cases yet")
    confirm_requirement(client, headers, rid2)
    gate = client.get(f"/v1/projects/{pid}/gate", headers=headers).json()
    assert gate["pass"] is False and gate["coverage_pct"] == 50.0
    breach = next(b for b in gate["breaches"] if b["check"] == "min_coverage")
    assert breach["limit"] == 80 and breach["actual"] == 50.0

    # a lenient threshold passes again; ?exit=1 on the failing default -> 412
    gate = client.get(f"/v1/projects/{pid}/gate", params={"min_coverage": 40},
                      headers=headers).json()
    assert gate["pass"] is True
    r = client.get(f"/v1/projects/{pid}/gate", params={"exit": 1}, headers=headers)
    assert r.status_code == 412


# ------------------------------------------------------------------ public API surface

def test_api_key_launches_and_reads_runs(client, register_org, create_project):
    headers = register_org()
    pid = create_project(headers)
    seed_approved_cases(client, headers, pid)
    eid = make_env(client, headers, pid)
    key = make_api_key(client, headers)["key"]
    key_headers = {"X-API-Key": key}

    # traceability readable via API key
    r = client.get(f"/v1/projects/{pid}/traceability", headers=key_headers)
    assert r.status_code == 200 and r.json()["rows"]

    # run launch via API key (public CI surface)
    r = client.post(f"/v1/projects/{pid}/runs", json={"environment_id": eid},
                    headers=key_headers)
    assert r.status_code == 202, r.text
    run_id = r.json()["run_id"]

    # run readable via API key; environment is dead -> all cases errored
    run = wait_run_terminal(client, key_headers, run_id)
    assert run["state"] == "completed"
    assert run["counts"]["errored"] == run["counts"]["total"] > 0

    # gate: errored cases are open defects -> max_failed=0 breach names REQ-001
    gate = client.get(f"/v1/projects/{pid}/gate", params={"max_failed": 0},
                      headers=key_headers).json()
    assert gate["pass"] is False
    assert gate["latest_run"]["id"] == run_id
    breach = next(b for b in gate["breaches"] if b["check"] == "max_failed")
    assert breach["actual"] > 0
    assert "REQ-001" in breach.get("requirement_external_ids", [])

    # no auth at all -> 401
    assert client.get(f"/v1/runs/{run_id}").status_code == 401


def test_api_key_is_org_scoped(client, register_org, create_project):
    headers_a = register_org("Org A")
    headers_b = register_org("Org B")
    pid_b = create_project(headers_b)
    key_a = make_api_key(client, headers_a)["key"]

    # org A's key cannot see org B's project
    r = client.get(f"/v1/projects/{pid_b}/gate", headers={"X-API-Key": key_a})
    assert r.status_code == 404
    r = client.get(f"/v1/projects/{pid_b}/traceability", headers={"X-API-Key": key_a})
    assert r.status_code == 404


# ------------------------------------------------------------------ exports (FR-070)

def test_xray_and_defects_exports(client, register_org, create_project):
    headers = register_org()
    pid = create_project(headers)
    seed_approved_cases(client, headers, pid)
    eid = make_env(client, headers, pid)
    r = client.post(f"/v1/projects/{pid}/runs", json={"environment_id": eid},
                    headers=headers)
    assert r.status_code == 202
    run_id = r.json()["run_id"]
    wait_run_terminal(client, headers, run_id)

    # Xray import JSON
    r = client.get(f"/v1/runs/{run_id}/exports/xray.json", headers=headers)
    assert r.status_code == 200
    assert "attachment" in r.headers.get("content-disposition", "")
    doc = json.loads(r.content)
    assert doc["info"]["summary"].startswith("Traceo run #")
    assert doc["tests"], "xray export has no tests"
    for t in doc["tests"]:
        assert t["status"] in ("PASSED", "FAILED")
        assert t["testInfo"]["type"] == "Generic" and t["testInfo"]["summary"]
    assert any(t.get("testKey") == "REQ-001" for t in doc["tests"])

    # Jira defects CSV: failures only, UTF-8 BOM so Excel reads it as UTF-8
    r = client.get(f"/v1/runs/{run_id}/exports/defects.csv", headers=headers)
    assert r.status_code == 200
    assert r.content.startswith(b"\xef\xbb\xbf"), "missing UTF-8 BOM"
    text = r.content.decode("utf-8-sig")
    assert text.splitlines()[0] == "Summary,Description,Priority,Labels"
    assert "REQ-001" in text and "[ERRORED]" in text


# ------------------------------------------------------------------ webhooks

def test_webhook_crud_test_and_fire(client, register_org, create_project, webhook_net):
    headers = register_org()
    pid = create_project(headers)

    # invalid scheme rejected
    r = client.post(f"/v1/projects/{pid}/webhooks",
                    json={"name": "bad", "url": "ftp://example.com/x"}, headers=headers)
    assert r.status_code == 422
    # unsupported event rejected
    r = client.post(f"/v1/projects/{pid}/webhooks",
                    json={"name": "bad", "url": "https://example.com/hook",
                          "events": ["run.deleted"]}, headers=headers)
    assert r.status_code == 422

    r = client.post(f"/v1/projects/{pid}/webhooks",
                    json={"name": "ci hook", "url": "https://example.com/hook",
                          "secret": "s3cret"}, headers=headers)
    assert r.status_code == 201, r.text
    hook = r.json()
    assert hook["secret_set"] is True and hook["events"] == ["run.completed"]

    hooks = client.get(f"/v1/projects/{pid}/webhooks", headers=headers).json()
    assert len(hooks) == 1 and "secret" not in hooks[0]

    r = client.patch(f"/v1/webhooks/{hook['id']}", json={"name": "renamed"},
                     headers=headers)
    assert r.status_code == 200 and r.json()["name"] == "renamed"

    # test-fire: delivered, signed with HMAC-SHA256 over the exact body
    r = client.post(f"/v1/webhooks/{hook['id']}/test", headers=headers)
    assert r.status_code == 200
    assert r.json() == {"webhook_id": hook["id"], "delivered": True, "status": 200}
    assert len(webhook_net) == 1
    call = webhook_net[0]
    body = call["content"]
    expected_sig = "sha256=" + hmac.new(b"s3cret", body, hashlib.sha256).hexdigest()
    assert call["headers"]["X-Traceo-Signature"] == expected_sig
    assert call["headers"]["X-Traceo-Event"] == "run.completed"
    payload = json.loads(body)
    assert payload["event"] == "run.completed" and payload["project"]["id"] == pid
    hooks = client.get(f"/v1/projects/{pid}/webhooks", headers=headers).json()
    assert hooks[0]["last_status"] == 200 and hooks[0]["last_fired_at"]

    # Slack special case: {"text": <summary line>} payload instead
    r = client.post(f"/v1/projects/{pid}/webhooks",
                    json={"name": "slack",
                          "url": "https://hooks.slack.com/services/T0/B0/XYZ"},
                    headers=headers)
    assert r.status_code == 201
    client.post(f"/v1/webhooks/{r.json()['id']}/test", headers=headers)
    slack_body = json.loads(webhook_net[-1]["content"])
    assert set(slack_body) == {"text"}
    assert slack_body["text"].startswith("Run #")
    assert "completed in project" in slack_body["text"]

    # fire_webhooks() delivers to every enabled subscribed hook and never raises
    from app.db import SessionLocal
    from app.modules.integrations import fire_webhooks
    db = SessionLocal()
    try:
        before = len(webhook_net)
        fire_webhooks(db, pid, "run.completed", {
            "event": "run.completed", "project": {"id": pid, "name": "p"},
            "run": {"id": "r1", "display_id": 1001, "state": "completed",
                    "counts": {"total": 2, "passed": 1, "failed": 1, "errored": 0}},
            "timestamp": "2026-01-01T00:00:00+00:00"})
        assert len(webhook_net) == before + 2  # json hook + slack hook
        # unsubscribed event fires nothing
        fire_webhooks(db, pid, "run.started", {"event": "run.started"})
        assert len(webhook_net) == before + 2
    finally:
        db.close()

    r = client.delete(f"/v1/webhooks/{hook['id']}", headers=headers)
    assert r.status_code == 200
    assert len(client.get(f"/v1/projects/{pid}/webhooks", headers=headers).json()) == 1


def test_webhook_fires_on_run_completion(client, register_org, create_project,
                                         webhook_net):
    headers = register_org()
    pid = create_project(headers)
    seed_approved_cases(client, headers, pid)
    eid = make_env(client, headers, pid)
    r = client.post(f"/v1/projects/{pid}/webhooks",
                    json={"name": "notify", "url": "https://example.com/done"},
                    headers=headers)
    assert r.status_code == 201

    r = client.post(f"/v1/projects/{pid}/runs", json={"environment_id": eid},
                    headers=headers)
    assert r.status_code == 202
    wait_run_terminal(client, headers, r.json()["run_id"])

    deadline = time.time() + 10
    while not webhook_net and time.time() < deadline:
        time.sleep(0.1)
    assert webhook_net, "run completion did not fire the webhook"
    payload = json.loads(webhook_net[-1]["content"])
    assert payload["event"] == "run.completed"
    assert payload["run"]["state"] == "completed"
    assert payload["run"]["counts"]["total"] > 0 and payload["timestamp"]


# ------------------------------------------------------------------ schedules

def test_schedule_crud_and_scheduler_tick(client, register_org, create_project):
    headers = register_org()
    pid = create_project(headers)
    seed_approved_cases(client, headers, pid)
    eid = make_env(client, headers, pid)

    # interval below 15 minutes rejected
    r = client.post(f"/v1/projects/{pid}/schedules",
                    json={"name": "too fast", "environment_id": eid,
                          "interval_minutes": 10}, headers=headers)
    assert r.status_code == 422

    r = client.post(f"/v1/projects/{pid}/schedules",
                    json={"name": "nightly", "environment_id": eid,
                          "interval_minutes": 60}, headers=headers)
    assert r.status_code == 201, r.text
    sched = r.json()
    assert sched["enabled"] is True and sched["next_run_at"]

    r = client.patch(f"/v1/schedules/{sched['id']}", json={"interval_minutes": 30},
                     headers=headers)
    assert r.status_code == 200 and r.json()["interval_minutes"] == 30

    # force the schedule due, then run one scheduler tick
    from app.db import SessionLocal
    from app.models import Schedule
    from app.modules.integrations import scheduler_tick
    db = SessionLocal()
    try:
        row = db.get(Schedule, sched["id"])
        row.next_run_at = datetime.now(timezone.utc) - timedelta(minutes=1)
        db.commit()
    finally:
        db.close()

    assert scheduler_tick() == 1
    runs = items_of(client.get(f"/v1/projects/{pid}/runs", headers=headers).json())
    assert len(runs) == 1, "scheduler did not launch a run"
    assert runs[0]["initiated_by"]  # schedule creator
    wait_run_terminal(client, headers, runs[0]["id"])

    # schedule advanced: last_run_at set, next_run_at back in the future
    updated = client.get(f"/v1/projects/{pid}/schedules", headers=headers).json()[0]
    assert updated["last_run_at"]
    assert updated["next_run_at"] > updated["last_run_at"]

    # audit trail recorded the scheduled launch
    audit = client.get("/v1/audit", headers=headers).json()
    actions = {e.get("action") for e in items_of(audit)}
    assert "run.scheduled" in actions

    # a second immediate tick does nothing (next_run_at is in the future)
    assert scheduler_tick() == 0

    r = client.delete(f"/v1/schedules/{sched['id']}", headers=headers)
    assert r.status_code == 200
    assert client.get(f"/v1/projects/{pid}/schedules", headers=headers).json() == []


# ------------------------------------------------------------------ report deliverables

def test_run_report_html_is_english_and_ltr(client, register_org, create_project):
    """The printable run report used to switch between an Arabic and an English
    label table and flip the document direction. There is one table now, and the
    document is always LTR English — no project field can change that."""
    headers = register_org()
    pid = create_project(headers, automation="manual")
    seed_approved_cases(client, headers, pid)
    eid = make_env(client, headers, pid)
    r = client.post(f"/v1/projects/{pid}/runs", json={"environment_id": eid},
                    headers=headers)
    assert r.status_code == 202, r.text
    run_id = r.json()["run_id"]
    wait_run_terminal(client, headers, run_id)

    r = client.get(f"/v1/runs/{run_id}/report.html", headers=headers)
    assert r.status_code == 200, r.text
    page = r.text
    assert '<html dir="ltr" lang="en">' in page
    assert 'dir="rtl"' not in page and 'lang="ar"' not in page
    assert "Run Report" in page and "Defect reports" in page and "All results" in page
    # no Arabic anywhere in the deliverable
    assert not any("\u0600" <= ch <= "\u06ff" for ch in page)
    # RTL-only CSS hacks are gone with the branch that needed them
    assert "border-inline-start" not in page and "text-align: start" not in page


def test_matrix_xlsx_sheets_are_ltr(client, register_org, create_project):
    """FR-RPT-07's RTL sheet flag was driven by project.language; with the column
    gone every sheet must come out left-to-right."""
    from io import BytesIO

    from openpyxl import load_workbook

    headers = register_org()
    pid = create_project(headers, automation="manual")
    seed_approved_cases(client, headers, pid)
    r = client.get(f"/v1/projects/{pid}/exports/matrix.xlsx", headers=headers)
    assert r.status_code == 200, r.text
    wb = load_workbook(BytesIO(r.content))
    assert wb.sheetnames == ["Requirements", "Test Cases", "Matrix", "Latest Results"]
    for ws in wb.worksheets:
        assert not ws.sheet_view.rightToLeft, ws.title


# ------------------------------------------------------------------ org export + reference

def test_organisation_export(client, register_org, create_project):
    headers = register_org()
    pid = create_project(headers)
    rid = add_requirement(client, headers, pid, "REQ-001", "A requirement to export")
    confirm_requirement(client, headers, rid)

    r = client.get("/v1/export/organisation", headers=headers)
    assert r.status_code == 200
    assert 'filename="traceo_export.json"' in r.headers.get("content-disposition", "")
    doc = json.loads(r.content)
    assert doc["organisation"]["name"]
    assert [p["id"] for p in doc["projects"]] == [pid]
    assert doc["requirements"][0]["external_id"] == "REQ-001"
    assert doc["audit_entry_count"] > 0
    assert "evidence" not in json.dumps(doc)  # PDPL: evidence excluded


def test_reference_catalog(client, register_org):
    headers = register_org()
    r = client.get("/v1/reference/features", headers=headers)
    assert r.status_code == 200
    data = r.json()
    features = data["features"]
    assert len(features) == 37 and len(data["groups"]) == 8
    assert len({f["id"] for f in features}) == 37
    groups = {g["key"] for g in data["groups"]}
    for f in features:
        assert f["id"].startswith("FR-") and f["group"] in groups
        assert f["priority"] in ("P0", "P1", "P2")
        assert f["status"] in ("built", "planned")
        assert f["name_en"] and f["description_en"]
        # English-only catalog: the Arabic twin fields are gone for good
        assert "name_ar" not in f and "description_ar" not in f
    by_id = {f["id"]: f for f in features}
    # addendum features are built; capabilities absent from this codebase are honest
    assert by_id["FR-061"]["status"] == "built"
    assert by_id["FR-060"]["status"] == "built"
    assert by_id["FR-070"]["status"] == "built"
    assert by_id["FR-011"]["status"] == "planned"
    assert by_id["FR-021"]["status"] == "planned"
    assert data["counts"]["built"] + data["counts"]["planned"] == 37
