"""Execution, generation and governance guarantees.

FR-040 per-run concurrency · FR-043 fixture lifecycle · FR-036 manual edits survive
regeneration · FR-033 auth-class negatives · FR-010 XLSX + paste ingestion ·
FR-071 bilingual export with gap and failure sheets · FR-082 audit retention/export.

The fixture tests drive a real HTTP server in-process: teardown-on-failure is the kind
of guarantee that only means something when something actually receives the DELETE.
"""
import json
import threading
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO

from conftest import (add_requirement, confirm_requirement, import_spec, items_of,
                      poll_job, small_openapi_spec)


# ---------------------------------------------------------------- fake SUT

class _Recorder(BaseHTTPRequestHandler):
    """Records every request; /orders creates a fixture, /boom always fails."""

    log: list = []
    delete_status = 204

    def _respond(self, status, payload):
        body = json.dumps(payload).encode()
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _read(self):
        length = int(self.headers.get("Content-Length") or 0)
        raw = self.rfile.read(length) if length else b""
        try:
            return json.loads(raw or b"{}")
        except ValueError:
            return {"_raw": raw.decode("utf-8", "replace")}

    def do_POST(self):
        body = self._read()
        type(self).log.append(("POST", self.path, body))
        if self.path.startswith("/orders"):
            self._respond(201, {"id": "fixture-1", "name": body.get("name", "")})
        else:
            self._respond(201, {"id": "obj-1", "name": body.get("name", "")})

    def do_GET(self):
        type(self).log.append(("GET", self.path, None))
        self._respond(200, {"id": "obj-1", "name": "ok"})

    def do_DELETE(self):
        type(self).log.append(("DELETE", self.path, None))
        self._respond(type(self).delete_status, {})

    def log_message(self, *_args):  # silence the default stderr logging
        pass


def _start_sut():
    _Recorder.log = []
    _Recorder.delete_status = 204
    server = ThreadingHTTPServer(("127.0.0.1", 0), _Recorder)
    threading.Thread(target=server.serve_forever, daemon=True).start()
    return server, f"http://127.0.0.1:{server.server_address[1]}"


def _project_with_case(client, headers, create_project, path="/customers",
                       expect_status=201):
    pid = create_project(headers, name="Lifecycle", language="en")
    import_spec(client, headers, pid)
    rid = add_requirement(client, headers, pid, "REQ-001", "Customers can be created",
                          criteria=["A valid customer is created"], priority="high")
    confirm_requirement(client, headers, rid)
    case = client.post(f"/v1/projects/{pid}/test-cases", json={
        "title": "Create customer", "type": "positive", "priority": "high",
        "steps": [{"method": "POST", "path": path,
                   "request": {"body": {"name": "A", "phone": "0512345678", "age": 30}},
                   "assertions": [{"type": "status_code", "expected": expect_status}]}],
        "requirement_ids": [rid],
    }, headers=headers).json()
    client.post(f"/v1/test-cases/{case['id']}/approve", headers=headers)
    return pid, rid, case["id"]


# ---------------------------------------------------------------- FR-043

FIXTURES = [{
    "name": "seed-order",
    "create": {"method": "POST", "path": "/orders",
               "body": {"name": "{{run_ns}}-order"}},
    "extract": {"order_id": "id"},
    "delete": {"method": "DELETE", "path": "/orders/{{order_id}}"},
}]


def test_fixtures_are_namespaced_created_and_torn_down(client, register_org,
                                                       create_project):
    server, base_url = _start_sut()
    try:
        headers = register_org("Fixture Org")
        pid, _rid, _case = _project_with_case(client, headers, create_project)
        env = client.post(f"/v1/projects/{pid}/environments", json={
            "name": "stg", "base_url": base_url, "fixtures": FIXTURES}, headers=headers).json()
        assert env["fixtures"], "fixtures must round-trip on the environment"

        r = client.post(f"/v1/projects/{pid}/runs",
                        json={"environment_id": env["id"]}, headers=headers)
        assert r.status_code == 202, r.text
        poll_job(client, headers, r.json()["job_id"])
        run = client.get(f"/v1/runs/{r.json()['run_id']}", headers=headers).json()

        methods = [(m, p) for m, p, _b in _Recorder.log]
        assert ("POST", "/orders") in methods                    # created before the suite
        assert ("DELETE", "/orders/fixture-1") in methods        # AC2 torn down after
        assert methods.index(("POST", "/orders")) < methods.index(("POST", "/customers"))
        assert methods.index(("DELETE", "/orders/fixture-1")) > methods.index(("POST", "/customers"))

        created_body = next(b for m, p, b in _Recorder.log if (m, p) == ("POST", "/orders"))
        assert created_body["name"].startswith("traceo-")        # AC1 run-namespaced

        assert run["fixtures"]["created"] == ["seed-order"]
        assert run["fixtures"]["removed"] == ["seed-order"]
        assert run["fixtures"]["orphans"] == []
    finally:
        server.shutdown()


def test_a_fixture_that_cannot_be_removed_is_reported(client, register_org,
                                                      create_project):
    server, base_url = _start_sut()
    try:
        headers = register_org("Orphan Org")
        pid, _rid, _case = _project_with_case(client, headers, create_project)
        env = client.post(f"/v1/projects/{pid}/environments", json={
            "name": "stg", "base_url": base_url, "fixtures": FIXTURES}, headers=headers).json()
        _Recorder.delete_status = 500     # teardown will fail

        r = client.post(f"/v1/projects/{pid}/runs",
                        json={"environment_id": env["id"]}, headers=headers)
        poll_job(client, headers, r.json()["job_id"])
        run = client.get(f"/v1/runs/{r.json()['run_id']}", headers=headers).json()

        orphans = run["fixtures"]["orphans"]
        assert len(orphans) == 1                                  # AC3
        assert orphans[0]["name"] == "seed-order"
        assert "500" in orphans[0]["reason"]
        assert run["fixtures"]["removed"] == []
    finally:
        server.shutdown()


def test_teardown_still_runs_when_every_case_fails(client, register_org, create_project):
    """AC2 — teardown is in a finally block, not on the happy path."""
    server, base_url = _start_sut()
    try:
        headers = register_org("Failing Org")
        # the SUT answers 201; the case demands 418, so every case in the run fails
        pid, _rid, _case = _project_with_case(client, headers, create_project,
                                              expect_status=418)
        env = client.post(f"/v1/projects/{pid}/environments", json={
            "name": "stg", "base_url": base_url, "fixtures": FIXTURES}, headers=headers).json()

        r = client.post(f"/v1/projects/{pid}/runs",
                        json={"environment_id": env["id"]}, headers=headers)
        poll_job(client, headers, r.json()["job_id"])
        run = client.get(f"/v1/runs/{r.json()['run_id']}", headers=headers).json()

        assert run["counts"]["failed"] >= 1, "this run is meant to fail"
        assert run["fixtures"]["removed"] == ["seed-order"]
    finally:
        server.shutdown()


# ---------------------------------------------------------------- FR-040

def test_per_run_concurrency_is_recorded_and_bounded(client, register_org, create_project):
    server, base_url = _start_sut()
    try:
        headers = register_org("Concurrency Org")
        pid, _rid, _case = _project_with_case(client, headers, create_project)
        env = client.post(f"/v1/projects/{pid}/environments",
                          json={"name": "stg", "base_url": base_url}, headers=headers).json()

        r = client.post(f"/v1/projects/{pid}/runs",
                        json={"environment_id": env["id"], "concurrency": 4,
                              "branch": "main"}, headers=headers)
        assert r.status_code == 202
        poll_job(client, headers, r.json()["job_id"])
        run = client.get(f"/v1/runs/{r.json()['run_id']}", headers=headers).json()
        assert run["branch"] == "main" and run["source"] == "manual"

        for bad in (0, 33):
            r = client.post(f"/v1/projects/{pid}/runs",
                            json={"environment_id": env["id"], "concurrency": bad},
                            headers=headers)
            assert r.status_code == 422, f"concurrency={bad} should be rejected"
            assert r.json()["detail"]["code"] == "invalid_concurrency"
    finally:
        server.shutdown()


# ---------------------------------------------------------------- FR-036 / FR-033

def test_regeneration_preserves_manual_edits_and_reports_changes(
        client, register_org, create_project):
    headers = register_org("Regen Org")
    pid = create_project(headers, name="Regen", language="en")
    import_spec(client, headers, pid)
    rid = add_requirement(client, headers, pid, "REQ-001",
                          "Create a customer with a valid phone and age",
                          criteria=["A valid customer is created"], priority="high")
    confirm_requirement(client, headers, rid)

    job = poll_job(client, headers, client.post(
        f"/v1/projects/{pid}/generate", json={"depth": "standard"},
        headers=headers).json()["job_id"])
    assert job["result"]["generated"] > 0
    assert job["result"]["preserved_manual_edits"] == 0
    assert job["result"]["changed_cases"], "the first run must report what it added"

    cases = items_of(client.get(f"/v1/projects/{pid}/test-cases", headers=headers).json())
    target = cases[0]
    edited = client.patch(f"/v1/test-cases/{target['id']}",
                          json={"title": target["title"],
                                "description": "Hand-tuned by the QA lead"},
                          headers=headers)
    assert edited.status_code == 200 and edited.json()["user_modified"] is True

    # Regenerate: the edited case must come back untouched.
    job2 = poll_job(client, headers, client.post(
        f"/v1/projects/{pid}/generate", json={"depth": "standard"},
        headers=headers).json()["job_id"])
    assert job2["result"]["preserved_manual_edits"] >= 1     # AC3

    after = client.get(f"/v1/test-cases/{target['id']}", headers=headers).json()
    assert after["description"] == "Hand-tuned by the QA lead"
    assert after["user_modified"] is True

    titles = [c["title"] for c in items_of(
        client.get(f"/v1/projects/{pid}/test-cases", headers=headers).json())]
    assert titles.count(target["title"]) == 1, "regeneration must not duplicate the case"


def test_generation_covers_the_full_auth_negative_class(client, register_org,
                                                        create_project):
    headers = register_org("Auth Cases Org")
    pid = create_project(headers, name="Auth", language="en")
    spec = small_openapi_spec()
    spec["paths"]["/customers"]["post"]["security"] = [{"bearerAuth": []}]
    spec["components"] = {"securitySchemes": {"bearerAuth": {"type": "http",
                                                             "scheme": "bearer"}}}
    import_spec(client, headers, pid, spec)
    rid = add_requirement(client, headers, pid, "REQ-001",
                          "Only an authorised user may create a customer",
                          criteria=["An unauthorised caller is rejected"], priority="high")
    confirm_requirement(client, headers, rid)
    poll_job(client, headers, client.post(f"/v1/projects/{pid}/generate",
                                          json={"depth": "standard"},
                                          headers=headers).json()["job_id"])

    titles = " | ".join(c["title"] for c in items_of(
        client.get(f"/v1/projects/{pid}/test-cases", headers=headers).json()))
    for expected in ("unauthenticated access", "expired credential",
                     "wrong-role credential", "malformed JSON body"):
        assert expected in titles, f"FR-033 AC1 requires a '{expected}' case: {titles}"


# ---------------------------------------------------------------- FR-010

def test_xlsx_requirements_document_is_accepted(client, register_org, create_project):
    try:
        from openpyxl import Workbook
    except ImportError:
        return
    headers = register_org("XLSX Org")
    pid = create_project(headers, name="Sheet", language="en")

    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "Requirement", "Priority"])
    ws.append(["REQ-100", "The system must reject an order with no items.", "high"])
    ws.append(["REQ-101", "The system must email a receipt after payment.", "medium"])
    buf = BytesIO()
    wb.save(buf)

    r = client.post(f"/v1/projects/{pid}/documents",
                    files={"file": ("reqs.xlsx", buf.getvalue(),
                                    "application/vnd.openxmlformats-officedocument."
                                    "spreadsheetml.sheet")},
                    headers=headers)
    assert r.status_code in (200, 201, 202), r.text
    poll_job(client, headers, r.json()["job_id"])
    ids = {q["external_id"] for q in items_of(
        client.get(f"/v1/projects/{pid}/requirements", headers=headers).json())}
    assert {"REQ-100", "REQ-101"} <= ids, f"XLSX rows must become requirements: {ids}"


def test_pasted_requirements_are_parsed(client, register_org, create_project):
    headers = register_org("Paste Org")
    pid = create_project(headers, name="Paste", language="ar")
    r = client.post(f"/v1/projects/{pid}/requirements/paste", json={
        "title": "ملاحظات الاجتماع",
        "text": "REQ-200: يجب أن يرفض النظام طلباً بدون عناصر.\n- رفض الطلب الفارغ برمز 422",
    }, headers=headers)
    assert r.status_code == 202, r.text
    poll_job(client, headers, r.json()["job_id"])
    ids = {q["external_id"] for q in items_of(
        client.get(f"/v1/projects/{pid}/requirements", headers=headers).json())}
    assert "REQ-200" in ids

    assert client.post(f"/v1/projects/{pid}/requirements/paste", json={"text": "  "},
                       headers=headers).status_code == 422


# ---------------------------------------------------------------- FR-071

def test_export_is_bilingual_and_carries_gap_and_failure_sheets(
        client, register_org, create_project):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return
    headers = register_org("Export Org")
    pid, rid, _case = _project_with_case(client, headers, create_project)
    # a second confirmed requirement with no case at all => a gap row
    gap_rid = add_requirement(client, headers, pid, "REQ-002", "Refunds are auditable")
    confirm_requirement(client, headers, gap_rid)

    r = client.get(f"/v1/projects/{pid}/exports/matrix.xlsx",
                   params={"lang": "both"}, headers=headers)
    assert r.status_code == 200, r.text[:300]
    wb = load_workbook(BytesIO(r.content))

    names = wb.sheetnames
    assert any("Gaps" in n for n in names), f"FR-071 AC1 needs a gap list: {names}"
    assert any("Failures" in n for n in names), f"FR-071 AC1 needs a failure list: {names}"
    assert any("المصفوفة" in n for n in names), f"bilingual sheet names expected: {names}"

    gaps = wb[next(n for n in names if "Gaps" in n)]
    header = [c.value for c in gaps[1]]
    assert any("/" in str(h) for h in header), "bilingual headers expected"
    gap_ids = [row[0] for row in gaps.iter_rows(min_row=2, values_only=True)]
    assert "REQ-002" in gap_ids and "REQ-001" not in gap_ids

    for sheet in wb.worksheets:
        assert sheet.oddFooter.left.text, "FR-071 AC4: every page carries the stamp"
        assert sheet.sheet_view.rightToLeft is True

    english = client.get(f"/v1/projects/{pid}/exports/matrix.xlsx",
                         params={"lang": "en"}, headers=headers)
    assert "Matrix" in load_workbook(BytesIO(english.content)).sheetnames


# ---------------------------------------------------------------- FR-082

def test_audit_retention_export_and_bounded_purge(client, register_org, create_project):
    headers = register_org("Audit Org")
    create_project(headers, name="Audited", language="en")

    entries = client.get("/v1/audit", headers=headers).json()["items"]
    assert entries and all(e["retain_until"] for e in entries)   # AC1 + stamped retention

    retention = client.get("/v1/audit/retention", headers=headers).json()
    assert retention["retention_days"] == 90                      # AC3 default
    assert retention["past_retention"] == 0

    r = client.put("/v1/audit/retention", json={"retention_days": 365}, headers=headers)
    assert r.status_code == 200
    assert client.get("/v1/audit/retention", headers=headers).json()["retention_days"] == 365
    assert client.put("/v1/audit/retention", json={"retention_days": 0},
                      headers=headers).status_code == 422

    # AC2 — purge cannot touch anything still inside its retention window
    before = len(client.get("/v1/audit", params={"limit": 200}, headers=headers).json()["items"])
    assert client.post("/v1/audit/purge", headers=headers).json()["removed"] == 0
    after = client.get("/v1/audit", params={"limit": 200}, headers=headers).json()["items"]
    assert len(after) >= before

    # AC4 — exportable
    export = client.get("/v1/audit/export.csv", headers=headers)
    assert export.status_code == 200
    assert "text/csv" in export.headers["content-type"]
    lines = export.text.strip().splitlines()
    assert lines[0].startswith("id,occurred_at,actor_id,action")
    assert len(lines) > 1


def test_audit_log_has_no_write_path(client, register_org):
    """AC2 — immutability is structural: no route mutates an entry."""
    headers = register_org("Immutable Org")
    entry_id = client.get("/v1/audit", headers=headers).json()["items"][0]["id"]
    for method in (client.patch, client.put, client.delete):
        r = method(f"/v1/audit/{entry_id}", headers=headers)
        assert r.status_code in (404, 405), \
            f"{method.__name__.upper()} /audit/{{id}} must not exist, got {r.status_code}"
