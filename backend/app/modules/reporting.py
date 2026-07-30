"""Reporting module (TRD §4.8) — the exportable deliverables.

Traceability matrix as styled XLSX (FR-RPT-04, RTL sheets for Arabic projects
FR-RPT-07), run reports as JSON + a self-contained printable HTML page that doubles
as the PDF deliverable via the browser's print dialog (FR-RPT-01/02/03/05), and
run-over-run regression comparison (FR-RPT-06).
"""
import html
import json
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy.orm import Session

from ..db import get_db
from ..deps import get_project_scoped, require
from ..models import (Environment, Project, Requirement, RequirementTestCase,
                      Run, TestCase, TestResult, User)
from .traceability import (GAP_NEXT_ACTIONS, derive_severity, gap_reason,
                           is_high_priority, run_display_id)

router = APIRouter()

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
EVIDENCE_HTML_MAX = 4000  # display cap; stored evidence is already truncated on capture


def _iso(dt: datetime | None) -> str | None:
    return dt.isoformat() if dt else None


def _get_run(run_id: str, user: User, db: Session) -> Run:
    run = db.get(Run, run_id)
    if not run or run.organisation_id != user.organisation_id:
        raise HTTPException(404, detail={"code": "not_found", "message": "Run not found"})
    return run


# ---------------------------------------------------------------------------
# Shared data assembly
# ---------------------------------------------------------------------------

def _requirements_by_case(db: Session, case_ids: list[str]) -> dict[str, list[dict]]:
    out: dict[str, list[dict]] = {cid: [] for cid in case_ids}
    if not case_ids:
        return out
    rows = (db.query(RequirementTestCase, Requirement)
            .join(Requirement, Requirement.id == RequirementTestCase.requirement_id)
            .filter(RequirementTestCase.test_case_id.in_(case_ids))
            .all())
    for link, req in rows:
        out[link.test_case_id].append({
            "id": req.id, "external_id": req.external_id, "description": req.description,
            "priority": req.priority,
        })
    return out


def _latest_result_map(db: Session, case_ids: list[str]) -> dict[str, TestResult]:
    """test_case_id -> most recent TestResult (created_at asc scan: last write wins)."""
    latest: dict[str, TestResult] = {}
    if not case_ids:
        return latest
    rows = (db.query(TestResult)
            .filter(TestResult.test_case_id.in_(case_ids))
            .order_by(TestResult.created_at.asc(), TestResult.id.asc())
            .all())
    for res in rows:
        latest[res.test_case_id] = res
    return latest


def _run_outcomes(db: Session, run_id: str) -> dict[str, TestResult]:
    """Latest result per case WITHIN one run (normally exactly one per case)."""
    latest: dict[str, TestResult] = {}
    rows = (db.query(TestResult).filter(TestResult.run_id == run_id)
            .order_by(TestResult.created_at.asc(), TestResult.id.asc()).all())
    for res in rows:
        latest[res.test_case_id] = res
    return latest


def _run_dict(run: Run) -> dict:
    return {
        "id": run.id, "project_id": run.project_id, "environment_id": run.environment_id,
        "state": run.state, "started_at": _iso(run.started_at),
        "finished_at": _iso(run.finished_at), "counts": run.counts or {},
        "initiated_by": run.initiated_by, "abort_reason": run.abort_reason,
        "created_at": _iso(run.created_at),
    }


def _report_entries(db: Session, run: Run) -> list[dict]:
    rows = (db.query(TestResult, TestCase)
            .join(TestCase, TestCase.id == TestResult.test_case_id)
            .filter(TestResult.run_id == run.id)
            .order_by(TestResult.created_at.asc())
            .all())
    reqs = _requirements_by_case(db, [tc.id for _res, tc in rows])
    entries = []
    for res, tc in rows:
        linked = reqs.get(tc.id, [])
        high = any(is_high_priority(r.get("priority")) for r in linked)
        entries.append({
            "test_case": {"id": tc.id, "title": tc.title, "description": tc.description,
                          "type": tc.type, "priority": tc.priority, "state": tc.state,
                          "technique": tc.technique},
            "test_case_version": res.test_case_version,
            "outcome": res.outcome,
            "duration_ms": res.duration_ms,
            "failure_reason": res.failure_reason,
            "evidence": res.evidence or [],
            "requirements": linked,
            "executed_at": _iso(res.created_at),
            # FR-052: severity only meaningful on failed/errored cases
            "severity": derive_severity(res.outcome, res.failure_reason, high)
            if res.outcome in ("failed", "errored") else None,
        })
    return entries


def _counts_of(run: Run, entries: list[dict]) -> dict:
    counts = dict(run.counts or {})
    if not counts:
        counts = {"total": len(entries), "passed": 0, "failed": 0, "errored": 0}
        for e in entries:
            counts[e["outcome"]] = counts.get(e["outcome"], 0) + 1
    return counts


# ---------------------------------------------------------------------------
# XLSX traceability matrix (FR-RPT-04, FR-RPT-07)
# ---------------------------------------------------------------------------

# Column and section labels, EN + AR. `lang=both` stamps "EN / AR" on every header
# so one file serves an English delivery review and an Arabic contractual annex
# without exporting twice (FR-071 AC3).
_XLSX_LABELS = {
    "Requirements": "المتطلبات", "Test Cases": "حالات الاختبار", "Matrix": "المصفوفة",
    "Gaps": "الفجوات", "Failures": "الإخفاقات", "Latest Results": "أحدث النتائج",
    "External ID": "المعرّف", "Description": "الوصف", "Type": "النوع",
    "Priority": "الأولوية", "State": "الحالة", "Version": "الإصدار",
    "Confidence": "الثقة", "Linked Cases": "الحالات المرتبطة", "Case ID": "معرّف الحالة",
    "Title": "العنوان", "Technique": "الأسلوب", "Source": "المصدر",
    "User Modified": "عُدّلت يدوياً", "Requirements ": "المتطلبات",
    "Requirement": "المتطلب", "Requirement Description": "وصف المتطلب",
    "Req State": "حالة المتطلب", "Case Title": "عنوان الحالة",
    "Case State": "حالة الحالة", "Latest Outcome": "أحدث نتيجة",
    "Outcome": "النتيجة", "Duration (ms)": "المدة (مللي ثانية)", "Run ID": "معرّف التشغيل",
    "Executed At": "وقت التنفيذ", "Reason": "السبب", "Next Action": "الإجراء التالي",
    "Severity": "الشدة", "Assertion": "التحقق", "Expected": "المتوقع",
    "Actual": "الفعلي", "Run": "التشغيل", "Environment": "البيئة", "Branch": "الفرع",
    "Exported": "تاريخ التصدير",
}

_GAP_REASONS_EN = {
    "no_reachable_endpoint": "No reachable endpoint — import a spec covering it, or link it by hand",
    "all_cases_disabled": "Linked cases exist but none is approved — approve one in review",
    "no_approved_cases": "No approved cases — generate cases for this requirement",
}


def _resolve_lang(requested: str | None, project: Project) -> str:
    lang = (requested or project.language or "en").lower()
    return lang if lang in ("en", "ar", "both") else "en"


def _label(text: str, lang: str) -> str:
    arabic = _XLSX_LABELS.get(text, text)
    if lang == "ar":
        return arabic
    if lang == "both" and arabic != text:
        return f"{text} / {arabic}"
    return text


@router.get("/projects/{project_id}/exports/matrix.xlsx")
def export_matrix(project_id: str, lang: str | None = None, run_id: str | None = None,
                  user: User = Depends(require("export")),
                  db: Session = Depends(get_db)):
    """FR-071 — matrix, failure list and gap list in one workbook. `lang` selects
    en | ar | both; `run_id` stamps and scopes the failure sheet to one run."""
    project = get_project_scoped(project_id, user, db)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, detail={
            "code": "missing_dependency",
            "message": "XLSX export requires the 'openpyxl' package (pip install openpyxl)."})

    lang = _resolve_lang(lang, project)
    rtl = lang in ("ar", "both")
    stamp_run = None
    if run_id:
        stamp_run = db.get(Run, run_id)
        if not stamp_run or stamp_run.organisation_id != user.organisation_id \
                or stamp_run.project_id != project_id:
            raise HTTPException(404, detail={"code": "not_found",
                                             "message": "Run not found in this project"})
    reqs = (db.query(Requirement)
            .filter(Requirement.project_id == project_id,
                    Requirement.organisation_id == user.organisation_id,
                    Requirement.state != "removed")
            .order_by(Requirement.external_id.asc(), Requirement.created_at.asc()).all())
    cases = (db.query(TestCase)
             .filter(TestCase.project_id == project_id,
                     TestCase.organisation_id == user.organisation_id)
             .order_by(TestCase.created_at.asc()).all())
    case_by_id = {c.id: c for c in cases}
    links = (db.query(RequirementTestCase)
             .filter(RequirementTestCase.test_case_id.in_(list(case_by_id))).all()
             if case_by_id else [])
    cases_by_req: dict[str, list[str]] = {}
    reqs_by_case: dict[str, list[str]] = {}
    req_by_id = {r.id: r for r in reqs}
    for link in links:
        if link.requirement_id in req_by_id and link.test_case_id in case_by_id:
            cases_by_req.setdefault(link.requirement_id, []).append(link.test_case_id)
            reqs_by_case.setdefault(link.test_case_id, []).append(link.requirement_id)
    latest = _latest_result_map(db, list(case_by_id))

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="FFFF8A22")

    wb = Workbook()
    wb.remove(wb.active)

    env_name = ""
    if stamp_run:
        env = db.get(Environment, stamp_run.environment_id)
        env_name = env.name if env else ""
    # FR-071 AC4: identity on every printed page, not just the first sheet.
    stamp = " · ".join(filter(None, [
        f"{_label('Run', lang)} #{run_display_id(db, stamp_run)}" if stamp_run else "",
        f"{_label('Environment', lang)}: {env_name}" if env_name else "",
        f"{_label('Branch', lang)}: {stamp_run.branch}" if stamp_run and stamp_run.branch else "",
        f"{_label('Exported', lang)}: {datetime.now().isoformat(timespec='seconds')}",
        f"Traceo · {project.name}",
    ]))

    def sheet(title: str, headers: list[str], widths: list[int]):
        # Excel forbids \ / * ? : [ ] in a sheet name, so the bilingual separator
        # differs from the one used inside cells.
        ws = wb.create_sheet(_label(title, lang).replace(" / ", " · ")[:31])
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col, value=_label(h, lang))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.freeze_panes = "A2"
        ws.sheet_view.rightToLeft = rtl  # FR-RPT-07: Arabic projects export RTL sheets
        ws.oddFooter.left.text = stamp[:255]
        ws.evenFooter.left.text = stamp[:255]
        return ws

    def req_label(rid: str) -> str:
        r = req_by_id.get(rid)
        return (r.external_id or r.id[:8]) if r else rid[:8]

    # -- Sheet 1: Requirements
    ws = sheet("Requirements",
               ["External ID", "Description", "Type", "Priority", "State",
                "Version", "Confidence", "Linked Cases"],
               [14, 70, 16, 10, 12, 9, 11, 13])
    for r in reqs:
        ws.append([r.external_id, r.description, r.type, r.priority, r.state,
                   r.version, round(r.confidence, 2), len(cases_by_req.get(r.id, []))])

    # -- Sheet 2: Test Cases
    ws = sheet("Test Cases",
               ["Case ID", "Title", "Type", "Priority", "State", "Technique",
                "Source", "User Modified", "Version", "Requirements"],
               [38, 60, 10, 10, 10, 14, 11, 13, 9, 24])
    for c in cases:
        ws.append([c.id, c.title, c.type, c.priority, c.state, c.technique,
                   "generated" if c.generated else "manual",
                   "yes" if c.user_modified else "no", c.version,
                   ", ".join(req_label(rid) for rid in reqs_by_case.get(c.id, []))])

    # -- Sheet 3: Matrix — one row per requirement<->case link; uncovered reqs still appear
    ws = sheet("Matrix",
               ["Requirement", "Requirement Description", "Req State", "Case ID",
                "Case Title", "Case State", "Latest Outcome"],
               [14, 55, 12, 38, 55, 10, 14])
    for r in reqs:
        linked = cases_by_req.get(r.id, [])
        if not linked:
            ws.append([r.external_id or r.id[:8], r.description, r.state,
                       "", "— NOT COVERED —", "", ""])
            continue
        for cid in linked:
            c = case_by_id[cid]
            res = latest.get(cid)
            ws.append([r.external_id or r.id[:8], r.description, r.state,
                       c.id, c.title, c.state, res.outcome if res else "not_run"])

    # -- Sheet 4: Gaps — every confirmed requirement without an approved case (FR-051)
    ws = sheet("Gaps",
               ["External ID", "Description", "Priority", "Reason", "Next Action"],
               [14, 60, 10, 46, 52])
    for r in reqs:
        if r.state != "confirmed":
            continue
        states = [case_by_id[cid].state for cid in cases_by_req.get(r.id, [])
                  if cid in case_by_id]
        if any(s == "approved" for s in states):
            continue
        reason = gap_reason(states)
        ws.append([r.external_id or r.id[:8], r.description, r.priority,
                   _GAP_REASONS_EN[reason] if lang != "ar" else GAP_NEXT_ACTIONS[reason],
                   GAP_NEXT_ACTIONS[reason] if lang != "en" else _GAP_REASONS_EN[reason]])

    # -- Sheet 5: Failures — the defect list (FR-052), scoped to one run when asked
    ws = sheet("Failures",
               ["Case ID", "Title", "Requirement", "Outcome", "Severity",
                "Assertion", "Expected", "Actual", "Run ID", "Executed At"],
               [38, 52, 14, 10, 11, 16, 30, 30, 38, 24])
    failure_source = (_run_outcomes(db, stamp_run.id) if stamp_run
                      else {cid: res for cid, res in latest.items()})
    for cid, res in sorted(failure_source.items(),
                           key=lambda kv: _iso(kv[1].created_at) or ""):
        if res.outcome not in ("failed", "errored") or cid not in case_by_id:
            continue
        c = case_by_id[cid]
        linked = reqs_by_case.get(cid, [])
        high = any(is_high_priority(req_by_id[rid].priority)
                   for rid in linked if rid in req_by_id)
        fr = res.failure_reason or {}
        assertion = fr.get("assertion") if isinstance(fr.get("assertion"), dict) else {}
        ws.append([c.id, c.title, ", ".join(req_label(rid) for rid in linked),
                   res.outcome, derive_severity(res.outcome, res.failure_reason, high),
                   assertion.get("type", "") or ("transport" if fr.get("error") else ""),
                   str(assertion.get("expected", assertion.get("value", "")))[:200],
                   str(fr.get("actual", fr.get("error", "")))[:200],
                   res.run_id, _iso(res.created_at)])

    # -- Sheet 6: Latest Results
    ws = sheet("Latest Results",
               ["Case ID", "Title", "Case State", "Outcome", "Duration (ms)",
                "Run ID", "Executed At"],
               [38, 60, 10, 10, 13, 38, 24])
    for c in cases:
        res = latest.get(c.id)
        ws.append([c.id, c.title, c.state,
                   res.outcome if res else "not_run",
                   res.duration_ms if res else "",
                   res.run_id if res else "",
                   _iso(res.created_at) if res else ""])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"traceo-matrix-{project.id[:8]}-{lang}.xlsx"
    return StreamingResponse(
        buf, media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# ---------------------------------------------------------------------------
# Run report — JSON (FR-RPT-01/02/03 + FR-044-lite perf block)
# ---------------------------------------------------------------------------

def _percentile(sorted_vals: list[int], q: float) -> int:
    """Nearest-rank percentile over a pre-sorted list."""
    idx = min(len(sorted_vals) - 1, max(0, round(q * (len(sorted_vals) - 1))))
    return sorted_vals[int(idx)]


def _perf_block(db: Session, run: Run) -> list[dict]:
    """Per-endpoint latency aggregation from evidence elapsed_ms (FR-044).
    Evidence entries are positional per step, so the step list names the endpoint."""
    rows = (db.query(TestResult, TestCase)
            .join(TestCase, TestCase.id == TestResult.test_case_id)
            .filter(TestResult.run_id == run.id)
            .all())
    buckets: dict[tuple[str, str], list[int]] = {}
    for res, tc in rows:
        steps = sorted(tc.steps, key=lambda s: s.order)
        for i, ev in enumerate(res.evidence or []):
            if not isinstance(ev, dict) or i >= len(steps):
                continue
            elapsed = ev.get("elapsed_ms")
            if not isinstance(elapsed, (int, float)):
                continue
            key = (steps[i].method.upper(), steps[i].path)
            buckets.setdefault(key, []).append(int(elapsed))
    perf = []
    for (method, path), vals in sorted(buckets.items(), key=lambda kv: (kv[0][1], kv[0][0])):
        vals.sort()
        perf.append({"method": method, "path": path,
                     "p50_ms": _percentile(vals, 0.50),
                     "p95_ms": _percentile(vals, 0.95),
                     "max_ms": vals[-1], "calls": len(vals)})
    return perf


@router.get("/runs/{run_id}/report")
def run_report(run_id: str, user: User = Depends(require("view")),
               db: Session = Depends(get_db)):
    run = _get_run(run_id, user, db)
    entries = _report_entries(db, run)
    run_payload = _run_dict(run)
    run_payload["display_id"] = run_display_id(db, run)
    return {"run": run_payload, "counts": _counts_of(run, entries), "cases": entries,
            "perf": _perf_block(db, run)}


# ---------------------------------------------------------------------------
# Run report — printable HTML (FR-RPT-05; the browser's print dialog makes the PDF)
# ---------------------------------------------------------------------------

_LABELS_EN = {
    "title": "Run Report", "project": "Project", "environment": "Environment",
    "run": "Run", "state": "State", "started": "Started", "finished": "Finished",
    "total": "Total", "passed": "Passed", "failed": "Failed", "errored": "Errored",
    "pass_rate": "Pass rate", "defects": "Defect reports", "results": "All results",
    "case": "Test case", "requirement": "Requirement", "steps": "Steps",
    "expected": "Expected", "actual": "Actual", "evidence": "Evidence",
    "request": "Request", "response": "Response", "assertion": "Assertion",
    "outcome": "Outcome", "duration": "Duration (ms)", "type": "Type",
    "priority": "Priority", "error": "Error",
    "no_failures": "No failures — every executed case passed.",
    "aborted": "Run aborted", "generated_by": "Generated by Traceo",
}
_LABELS_AR = {
    "title": "تقرير التشغيل", "project": "المشروع", "environment": "البيئة",
    "run": "التشغيل", "state": "الحالة", "started": "بدأ", "finished": "انتهى",
    "total": "الإجمالي", "passed": "ناجح", "failed": "فاشل", "errored": "خطأ",
    "pass_rate": "نسبة النجاح", "defects": "تقارير العيوب", "results": "جميع النتائج",
    "case": "حالة الاختبار", "requirement": "المتطلب", "steps": "الخطوات",
    "expected": "المتوقع", "actual": "الفعلي", "evidence": "الدليل",
    "request": "الطلب", "response": "الاستجابة", "assertion": "التحقق",
    "outcome": "النتيجة", "duration": "المدة (مللي ثانية)", "type": "النوع",
    "priority": "الأولوية", "error": "الخطأ",
    "no_failures": "لا توجد إخفاقات — نجحت جميع الحالات المنفّذة.",
    "aborted": "تم إجهاض التشغيل", "generated_by": "تم إنشاؤه بواسطة Traceo",
}

_REPORT_CSS = """
:root { --bg:#131217; --card:#1B1A21; --border:#312F3C; --amber:#FF8A22;
        --text:#EDEBF2; --muted:#9C98AB; --pass:#3DBB78; --fail:#E5534B; --err:#D9A03C; }
* { box-sizing: border-box; }
body { background: var(--bg); color: var(--text); margin: 0; padding: 32px;
       font: 14px/1.55 -apple-system, "Segoe UI", Tahoma, Arial, sans-serif; }
.mono { font-family: "SF Mono", ui-monospace, Menlo, Consolas, monospace; font-size: 12px; }
h1 { font-size: 24px; margin: 0 0 4px; }
h1 .accent { color: var(--amber); }
h2 { font-size: 17px; margin: 34px 0 12px; border-inline-start: 4px solid var(--amber);
     padding-inline-start: 10px; }
.meta { color: var(--muted); margin-bottom: 24px; }
.meta .mono { color: var(--text); }
.kpis { display: flex; flex-wrap: wrap; gap: 12px; margin: 20px 0 8px; }
.kpi { background: var(--card); border: 1px solid var(--border); border-radius: 10px;
       padding: 14px 20px; min-width: 128px; }
.kpi .num { font-size: 26px; font-weight: 700; }
.kpi .lbl { color: var(--muted); font-size: 12px; margin-top: 2px; }
.kpi.pass .num { color: var(--pass); } .kpi.fail .num { color: var(--fail); }
.kpi.err .num { color: var(--err); } .kpi.rate .num { color: var(--amber); }
.badge { display: inline-block; padding: 2px 10px; border-radius: 999px; font-size: 12px;
         font-weight: 600; border: 1px solid var(--border); }
.badge.passed { color: var(--pass); } .badge.failed { color: var(--fail); }
.badge.errored { color: var(--err); }
.card { background: var(--card); border: 1px solid var(--border); border-radius: 10px;
        padding: 18px 20px; margin-bottom: 16px; page-break-inside: avoid; }
.card h3 { margin: 0 0 8px; font-size: 15px; }
.reqs { color: var(--muted); font-size: 13px; margin-bottom: 10px; }
.reqs .rid { color: var(--amber); font-weight: 600; }
.kv { display: grid; grid-template-columns: 110px 1fr; gap: 4px 14px; margin: 10px 0; }
.kv .k { color: var(--muted); }
pre { background: var(--bg); border: 1px solid var(--border); border-radius: 8px;
      padding: 10px 12px; overflow-x: auto; white-space: pre-wrap; word-break: break-word;
      font-family: "SF Mono", ui-monospace, Menlo, Consolas, monospace; font-size: 12px;
      margin: 6px 0 12px; direction: ltr; text-align: left; }
.step { border-top: 1px dashed var(--border); padding-top: 10px; margin-top: 10px; }
.step .line { font-weight: 600; }
table { border-collapse: collapse; width: 100%; background: var(--card);
        border: 1px solid var(--border); border-radius: 10px; overflow: hidden; }
th { background: var(--amber); color: #131217; text-align: start; font-size: 12px;
     padding: 9px 12px; }
td { border-top: 1px solid var(--border); padding: 8px 12px; font-size: 13px;
     vertical-align: top; }
.footer { color: var(--muted); font-size: 12px; margin-top: 36px; }
/* FR-071 AC4 — run identity repeated on every printed page. */
.page-stamp { display: none; }
@media print {
  :root { --bg:#FFFFFF; --card:#FFFFFF; --border:#D8D5E0; --text:#1B1A21;
          --muted:#5E5A6E; }
  body { padding: 0; margin-bottom: 22mm; }
  .page-stamp { display: block; position: fixed; bottom: 0; inset-inline: 0;
                font-size: 10px; color: #5E5A6E; padding: 4px 0;
                border-top: 1px solid #D8D5E0; }
  .kpi, .card, table { box-shadow: none; }
  th { color: #FFFFFF; background: #FF8A22; -webkit-print-color-adjust: exact;
       print-color-adjust: exact; }
}
"""


def _esc(v) -> str:
    return html.escape("" if v is None else str(v), quote=True)


def _jsonish(v) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v
    try:
        return json.dumps(v, ensure_ascii=False, indent=2, default=str)
    except Exception:  # noqa: BLE001
        return str(v)


def _clip(text: str) -> str:
    if len(text) > EVIDENCE_HTML_MAX:
        return text[:EVIDENCE_HTML_MAX] + "…[truncated]"
    return text


def _defect_card(entry: dict, L: dict) -> str:
    tc = entry["test_case"]
    fr = entry["failure_reason"] or {}
    reqs = "".join(
        f'<div><span class="rid mono">{_esc(r["external_id"] or r["id"][:8])}</span> '
        f'{_esc(r["description"])}</div>'
        for r in entry["requirements"])
    outcome_lbl = L.get(entry["outcome"], entry["outcome"])

    diag = ""
    if fr.get("assertion") is not None:
        diag = (f'<div class="kv">'
                f'<div class="k">{L["assertion"]}</div>'
                f'<div class="mono">{_esc(_jsonish(fr.get("assertion")))}</div>'
                f'<div class="k">{L["expected"]}</div>'
                f'<div class="mono">{_esc(_jsonish(fr.get("expected")))}</div>'
                f'<div class="k">{L["actual"]}</div>'
                f'<div class="mono">{_esc(_jsonish(fr.get("actual")))}</div>'
                f'</div>')
    elif fr.get("error"):
        diag = (f'<div class="kv"><div class="k">{L["error"]}</div>'
                f'<div class="mono">{_esc(fr["error"])}</div></div>')

    steps_html = []
    for i, ev in enumerate(entry["evidence"]):
        req_ev = ev.get("request") or {}
        resp_ev = ev.get("response")
        block = [f'<div class="step"><div class="line mono">'
                 f'{i + 1}. {_esc(req_ev.get("method", ""))} {_esc(req_ev.get("url", ""))}'
                 f' — {ev.get("elapsed_ms", 0)}ms</div>']
        req_parts = {k: req_ev.get(k) for k in ("headers", "body") if req_ev.get(k)}
        if req_parts:
            block.append(f'<div class="k">{L["request"]}</div>'
                         f'<pre>{_esc(_clip(_jsonish(req_parts)))}</pre>')
        if resp_ev is not None:
            block.append(
                f'<div class="k">{L["response"]} — HTTP {_esc(resp_ev.get("status"))}</div>'
                f'<pre>{_esc(_clip(_jsonish(resp_ev.get("body"))))}</pre>')
        elif ev.get("error"):
            block.append(f'<div class="k">{L["error"]}</div>'
                         f'<pre>{_esc(_clip(str(ev["error"])))}</pre>')
        block.append('</div>')
        steps_html.append("".join(block))

    return (f'<div class="card">'
            f'<h3>{_esc(tc["title"])} '
            f'<span class="badge {entry["outcome"]}">{_esc(outcome_lbl)}</span></h3>'
            f'<div class="reqs"><strong>{L["requirement"]}:</strong>{reqs or " —"}</div>'
            f'{diag}'
            f'<div class="k" style="color:var(--muted)">'
            f'{L["steps"]} / {L["evidence"]}</div>'
            f'{"".join(steps_html) or "<div class=step>—</div>"}'
            f'</div>')


@router.get("/runs/{run_id}/report.html", response_class=HTMLResponse)
def run_report_html(run_id: str, lang: str | None = None,
                    user: User = Depends(require("view")),
                    db: Session = Depends(get_db)):
    """The PDF deliverable (print this page). `lang` selects en | ar | both —
    bilingual renders every label as "EN / AR" in one document (FR-071 AC3)."""
    run = _get_run(run_id, user, db)
    project = db.get(Project, run.project_id)
    env = db.get(Environment, run.environment_id)
    entries = _report_entries(db, run)
    counts = _counts_of(run, entries)
    display_id = run_display_id(db, run)

    chosen = _resolve_lang(lang, project) if project else (lang or "en")
    arabic = chosen == "ar"
    if chosen == "both":
        L = {k: (f"{v} / {_LABELS_AR[k]}" if _LABELS_AR.get(k) and _LABELS_AR[k] != v else v)
             for k, v in _LABELS_EN.items()}
    else:
        L = _LABELS_AR if arabic else _LABELS_EN
    dir_attr, lang_attr = ("rtl", "ar") if arabic else ("ltr", "en")

    total = counts.get("total", 0)
    passed = counts.get("passed", 0)
    rate = f"{(passed / total * 100):.1f}%" if total else "—"

    failures = [e for e in entries if e["outcome"] in ("failed", "errored")]
    defects = "".join(_defect_card(e, L) for e in failures) \
        or f'<div class="card">{L["no_failures"]}</div>'

    result_rows = "".join(
        f'<tr><td>{_esc(e["test_case"]["title"])}'
        f'<div class="mono" style="color:var(--muted)">{_esc(e["test_case"]["id"])}</div></td>'
        f'<td>{_esc(e["test_case"]["type"])}</td>'
        f'<td>{_esc(e["test_case"]["priority"])}</td>'
        f'<td><span class="badge {e["outcome"]}">{_esc(L.get(e["outcome"], e["outcome"]))}</span></td>'
        f'<td class="mono">{e["duration_ms"]}</td>'
        f'<td>{"".join(f"<div class=mono>{_esc(r['external_id'] or r['id'][:8])}</div>" for r in e["requirements"]) or "—"}</td>'
        f'</tr>'
        for e in entries)

    abort_html = ""
    if run.state == "aborted":
        abort_html = (f'<div class="card"><h3>{L["aborted"]}</h3>'
                      f'<pre>{_esc(run.abort_reason or "")}</pre></div>')

    page = f"""<!DOCTYPE html>
<html dir="{dir_attr}" lang="{lang_attr}">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{L["title"]} — {_esc(project.name if project else run.project_id)}</title>
<style>{_REPORT_CSS}</style>
</head>
<body>
<h1><span class="accent">Traceo</span> — {L["title"]}</h1>
<div class="meta">
  {L["project"]}: <strong>{_esc(project.name if project else "?")}</strong> ·
  {L["environment"]}: <strong>{_esc(env.name if env else "?")}</strong> ·
  {L["run"]}: <span class="mono">#{display_id}</span>
  <span class="mono" style="color:var(--muted)">{_esc(run.id)}</span><br>
  {L["state"]}: {_esc(run.state)} ·
  {L["started"]}: <span class="mono">{_esc(_iso(run.started_at) or "—")}</span> ·
  {L["finished"]}: <span class="mono">{_esc(_iso(run.finished_at) or "—")}</span>
</div>
{abort_html}
<div class="kpis">
  <div class="kpi"><div class="num">{total}</div><div class="lbl">{L["total"]}</div></div>
  <div class="kpi pass"><div class="num">{passed}</div><div class="lbl">{L["passed"]}</div></div>
  <div class="kpi fail"><div class="num">{counts.get("failed", 0)}</div><div class="lbl">{L["failed"]}</div></div>
  <div class="kpi err"><div class="num">{counts.get("errored", 0)}</div><div class="lbl">{L["errored"]}</div></div>
  <div class="kpi rate"><div class="num">{rate}</div><div class="lbl">{L["pass_rate"]}</div></div>
</div>
<h2>{L["defects"]}</h2>
{defects}
<h2>{L["results"]}</h2>
<table>
<thead><tr><th>{L["case"]}</th><th>{L["type"]}</th><th>{L["priority"]}</th>
<th>{L["outcome"]}</th><th>{L["duration"]}</th><th>{L["requirement"]}</th></tr></thead>
<tbody>{result_rows or '<tr><td colspan="6">—</td></tr>'}</tbody>
</table>
<div class="footer">{L["generated_by"]} · {_esc(_iso(run.created_at) or "")}</div>
<div class="page-stamp">{L["run"]} #{display_id} · {_esc(env.name if env else "")}
{f" · {_esc(run.branch)}" if run.branch else ""} · {_esc(_iso(run.created_at) or "")}</div>
</body>
</html>"""
    return HTMLResponse(content=page)


# ---------------------------------------------------------------------------
# Run comparison (FR-RPT-06)
# ---------------------------------------------------------------------------

@router.get("/runs/{run_id}/compare/{other_id}")
def compare_runs(run_id: str, other_id: str, user: User = Depends(require("view")),
                 db: Session = Depends(get_db)):
    run = _get_run(run_id, user, db)
    other = _get_run(other_id, user, db)
    if run.project_id != other.project_id:
        raise HTTPException(409, detail={
            "code": "different_projects",
            "message": "Runs belong to different projects and cannot be compared"})

    current = _run_outcomes(db, run.id)
    baseline = _run_outcomes(db, other.id)
    shared = set(current) & set(baseline)
    titles = {}
    if shared:
        titles = {tc.id: tc.title for tc in db.query(TestCase)
                  .filter(TestCase.id.in_(list(shared))).all()}

    newly_failing, newly_passing = [], []
    unchanged = 0
    for cid in sorted(shared):
        now_o, prev_o = current[cid].outcome, baseline[cid].outcome
        item = {"test_case_id": cid, "title": titles.get(cid, ""),
                "outcome": now_o, "previous_outcome": prev_o}
        if now_o in ("failed", "errored") and prev_o == "passed":
            newly_failing.append(item)
        elif now_o == "passed" and prev_o in ("failed", "errored"):
            newly_passing.append(item)
        elif now_o == prev_o:
            unchanged += 1

    def _coverage(r: Run) -> float:
        c = r.counts or {}
        total = c.get("total", 0)
        return round(c.get("passed", 0) / total * 100, 1) if total else 0.0

    return {"run_id": run.id, "other_id": other.id,
            "newly_failing": newly_failing, "newly_passing": newly_passing,
            "unchanged": unchanged,
            "coverage_delta": round(_coverage(run) - _coverage(other), 1)}
